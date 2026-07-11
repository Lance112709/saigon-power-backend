"""Dispute center: package building, draft rendering, outcome allocation."""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from fakedb import FakeDB
from app.services.disputes import (
    _build_xlsx, _claims_from_cases, _claims_from_finding, _draft_email,
    build_dispute_package, record_outcome,
)

SUP = "00000000-0000-0000-0000-000000000001"
E1 = "1008901000000000000001"
E2 = "1008901000000000000002"


def case(esiid, loss, month="2026-04-01", issue="short_paid", name="Customer"):
    return {"supplier_id": SUP, "billing_month": month, "esiid": esiid,
            "issue_type": issue, "workflow_status": "open",
            "estimated_loss": loss, "recovered_amount": 0,
            "customer_name": name, "explanation": "wrong rate", "priority_score": 1}


def _seed(db, cases):
    db.tables["suppliers"] = [{"id": SUP, "name": "Discount Power",
                               "contact_email": "rep@discountpower.com"}]
    for c in cases:
        db.table("exception_cases").insert(c).execute()


def test_draft_email_totals_and_tone():
    claims = _claims_from_cases([case(E1, 3.0), case(E2, 5.5)])
    subject, body = _draft_email("Discount Power", claims, None, ["2026-04"], 8.5)
    assert "$8.50" in subject and "2026-04" in subject and "319010" in subject
    assert "2 account(s)" in body
    assert "true-up" in body


def test_xlsx_contains_all_accounts_and_total():
    import openpyxl
    blob = _build_xlsx("Discount Power", _claims_from_cases([case(E1, 3.0), case(E2, 5.5)]))
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert wb.sheetnames == ["Summary", "Accounts"]
    summary = list(wb["Summary"].iter_rows(values_only=True))
    assert summary[1][summary[0].index("Total Claimed $")] == 8.5
    accounts = list(wb["Accounts"].iter_rows(values_only=True))
    assert len(accounts) == 3  # header + 2 rows
    esiids = {r[0] for r in accounts[1:]}
    assert esiids == {E1, E2}


def test_finding_claims_use_per_account_losses_even_without_cases():
    """The May 2026 Discount Power bug: cases missing/$0, but the finding's
    account breakdown carries the real per-account losses."""
    db = FakeDB()  # no exception_cases at all
    finding = {
        "id": "f1", "supplier_id": SUP, "billing_month": "2026-05-01",
        "finding_type": "systemic_rate_change",
        "details": {"rate_from": 0.008, "rate_to": 0.005, "accounts": [
            {"esiid": E1, "rate_from": 0.008, "rate_to": 0.005, "kwh": 1000.0,
             "monthly_loss": 3.0},
            {"esiid": E2, "rate_from": 0.008, "rate_to": 0.005, "kwh": 2000.0,
             "monthly_loss": 6.0},
        ]},
    }
    claims = _claims_from_finding(db, finding)
    assert len(claims) == 2
    assert sum(c["claimed"] for c in claims) == 9.0
    assert claims[0]["case_id"] is None
    assert "0.008" in claims[0]["explanation"]


def test_build_dispute_from_finding_claims_real_total():
    db = FakeDB()
    db.tables["suppliers"] = [{"id": SUP, "name": "Discount Power",
                               "contact_email": "rep@discountpower.com"}]
    db.tables["audit_findings"] = [{
        "id": "f1", "supplier_id": SUP, "billing_month": "2026-05-01",
        "finding_type": "systemic_rate_change", "status": "open",
        "title": "Discount Power cut 2 accounts from 0.008 to 0.005 $/kWh",
        "explanation": "why text",
        "details": {"accounts": [
            {"esiid": E1, "rate_from": 0.008, "rate_to": 0.005, "kwh": 1000.0,
             "monthly_loss": 3.0},
            {"esiid": E2, "rate_from": 0.008, "rate_to": 0.005, "kwh": 2000.0,
             "monthly_loss": 6.0},
        ]},
    }]
    d = build_dispute_package(db, SUP, "lance", finding_id="f1")
    assert d["total_claimed"] == 9.0
    assert d["items_count"] == 2
    amounts = sorted(i["claimed_amount"] for i in db.tables["dispute_items"])
    assert amounts == [3.0, 6.0]


def test_build_dispute_package_creates_draft_and_links_cases():
    db = FakeDB()
    _seed(db, [case(E1, 3.0), case(E2, 5.5)])
    case_ids = [c["id"] for c in db.tables["exception_cases"]]
    d = build_dispute_package(db, SUP, "lance", case_ids=case_ids)
    assert d["status"] == "draft"
    assert d["total_claimed"] == 8.5
    assert d["email_to"] == "rep@discountpower.com"
    assert d["items_count"] == 2
    assert d["attachment_path"].startswith("statements/disputes/")
    # cases linked
    assert all(c["dispute_id"] == d["id"] for c in db.tables["exception_cases"])
    # attachment actually stored
    assert any(k.endswith(".xlsx") for k in db.storage.blobs)


def test_build_dispute_requires_cases():
    db = FakeDB()
    _seed(db, [])
    try:
        build_dispute_package(db, SUP, "lance", case_ids=["nope"])
        assert False, "should have raised"
    except ValueError:
        pass


def test_recovered_outcome_allocates_proportionally():
    db = FakeDB()
    _seed(db, [case(E1, 3.0), case(E2, 9.0)])
    case_ids = [c["id"] for c in db.tables["exception_cases"]]
    d = build_dispute_package(db, SUP, "lance", case_ids=case_ids)

    record_outcome(db, d["id"], "recovered", 6.0, "true-up on May stmt", "lance")

    dispute = db.tables["disputes"][0]
    assert dispute["status"] == "recovered"
    assert dispute["total_recovered"] == 6.0
    by_esiid = {i["esiid"]: i for i in db.tables["dispute_items"]}
    assert by_esiid[E1]["recovered_amount"] == 1.5   # 3/12 of $6
    assert by_esiid[E2]["recovered_amount"] == 4.5   # 9/12 of $6
    cases = {c["esiid"]: c for c in db.tables["exception_cases"]}
    assert cases[E1]["workflow_status"] == "recovered"
    assert cases[E1]["recovered_amount"] == 1.5


def test_rejected_outcome_records_no_recovery():
    db = FakeDB()
    _seed(db, [case(E1, 3.0)])
    d = build_dispute_package(db, SUP, "lance",
                              case_ids=[db.tables["exception_cases"][0]["id"]])
    record_outcome(db, d["id"], "rejected", 0, "provider says contract allows", "lance")
    assert db.tables["disputes"][0]["status"] == "rejected"
    assert db.tables["exception_cases"][0]["workflow_status"] == "open"
