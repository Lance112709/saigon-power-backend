from fastapi import APIRouter, HTTPException, Query, Body, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional
import re
import io
from datetime import datetime as dt
from app.db.client import get_client
from app.auth.deps import get_current_user, require_admin, require_manager, UserContext

router = APIRouter()

# Provider name → supplier code
PROVIDER_TO_CODE = {
    "BUDGET POWER": "BUDGET",
    "IRON HORSE": "IRONHORSE",
    "HERITAGE POWER": "HERITAGE",
    "NRG ENERGY": "NRG_COMM",
    "DISCOUNT POWER": "NRG",
    "CHARIOT ENERGY": "CHARIOT",
    "CLEANSKY ENERGY": "CLEANSKY",
    "HUDSON ENERGY": "HUDSON",
}

# Flexible supplier name → code (lowercase keys)
SUPPLIER_ALIASES: dict = {
    "budget": "BUDGET", "budget power": "BUDGET",
    "iron horse": "IRONHORSE", "ironhorse": "IRONHORSE",
    "heritage": "HERITAGE", "heritage power": "HERITAGE",
    "nrg": "NRG", "nrg energy": "NRG", "discount power": "NRG",
    "chariot": "CHARIOT", "chariot energy": "CHARIOT",
    "cleansky": "CLEANSKY", "cleansky energy": "CLEANSKY",
    "hudson": "HUDSON", "hudson energy": "HUDSON",
    "reliant": "RELIANT",
    "direct energy": "DIRECT", "direct": "DIRECT",
    "cirro": "CIRRO", "cirro energy": "CIRRO",
    "true power": "TRUEPOWER", "truepower": "TRUEPOWER",
    "nrg commercial": "NRG_COMM",
    "value power": "VALUEPOWER", "value": "VALUEPOWER",
    "tara energy": "TARA", "tara": "TARA",
    "pulse power": "PULSE", "pulse": "PULSE",
    "apg&e": "APGE", "apge": "APGE",
    "power next": "POWERNEXT", "powernext": "POWERNEXT",
    "goodcharlie": "GOODCHARLIE", "good charlie": "GOODCHARLIE",
    "sfe energy": "SFE", "sfe": "SFE",
}

ACTIVE_STATUSES  = {"active", "yes", "open", "current", "enrolled", "a", "1"}
INACTIVE_STATUSES = {"inactive", "no", "closed", "cancelled", "canceled",
                     "churned", "terminated", "dropped", "drop", "cancel", "i", "0"}

TEMPLATE_HEADERS = [
    "First Name", "Last Name", "Email", "Phone", "Date of Birth",
    "Mailing Address", "City", "State", "Zip Code",
    "ESIID", "Supplier", "Deal Status", "Sales Agent", "Service Address",
    "Contract Start Date", "Contract End Date", "Contract Signed Date",
    "Term (Months)", "Energy Rate", "Adder", "Product Type", "Meter Type",
    "ANXH", "Business Name", "Deal Owner",
]

SAMPLE_ROW = [
    "John", "Smith", "john.smith@gmail.com", "5551234567", "01/15/1985",
    "123 Main St", "Houston", "TX", "77001",
    "10089012345678901", "Budget Power", "ACTIVE", "Lance Nguyen", "123 Main St",
    "01/01/2024", "01/01/2026", "12/15/2023",
    "24", "0.089", "0.0005", "Fixed Rate", "Residential",
    "", "", "",
]

VALID_SUPPLIERS = [
    "Budget Power", "Iron Horse", "Heritage Power", "NRG Energy",
    "Discount Power", "Chariot Energy", "CleanSky Energy", "Hudson Energy",
    "Reliant", "Direct Energy", "Cirro Energy", "True Power",
]

def _extract_email(contact_str: str) -> Optional[str]:
    if not contact_str:
        return None
    m = re.search(r"\(([^)@]+@[^)]+)\)", str(contact_str))
    return m.group(1).strip().lower() if m else None

def _to_date_str(val) -> Optional[str]:
    if not val or str(val).strip() in ("", "nan", "None", "NaT"):
        return None
    try:
        from datetime import datetime as dt
        if hasattr(val, "date"):
            return val.date().isoformat()
        s = str(val).strip()
        for fmt in ("%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return dt.strptime(s[:16] if ":" in s else s[:10], fmt).date().isoformat()
            except Exception:
                continue
    except Exception:
        pass
    return None

def _to_float(val) -> Optional[float]:
    try:
        return float(str(val).replace(",", "").strip()) if val not in (None, "", "nan") else None
    except Exception:
        return None

def _enrich_deals(db, deals: list) -> None:
    """Backfill energy_rate and sales_agent from actual_commissions by ESIID."""
    missing_rate  = [d["esiid"] for d in deals if not d.get("energy_rate") and d.get("esiid")]
    missing_agent = [d["esiid"] for d in deals if not d.get("sales_agent") and d.get("esiid")]
    esiids_needed = list(set(missing_rate + missing_agent))
    if not esiids_needed:
        return
    comm_res = db.table("actual_commissions").select(
        "raw_esiid, raw_rate, raw_row_data, billing_month"
    ).in_("raw_esiid", esiids_needed).order("billing_month", desc=True).limit(len(esiids_needed) * 3).execute()
    latest: dict = {}
    for row in (comm_res.data or []):
        esiid = row.get("raw_esiid")
        if esiid and esiid not in latest:
            latest[esiid] = row
    for deal in deals:
        esiid = deal.get("esiid")
        if esiid and esiid in latest:
            comm = latest[esiid]
            if not deal.get("energy_rate") and comm.get("raw_rate"):
                deal["energy_rate"] = comm["raw_rate"]
            if not deal.get("sales_agent"):
                raw = comm.get("raw_row_data") or {}
                for key in ("AE Name", "Agent", "Sales Rep", "Rep Name", "Salesperson", "sales_agent"):
                    val = raw.get(key)
                    if val and str(val).strip():
                        deal["sales_agent"] = str(val).strip()
                        break

# ── Stats ─────────────────────────────────────────────────────────────────────

@router.get("/stats")
def get_crm_stats():
    db = get_client()
    customers = db.table("crm_customers").select("id", count="exact").execute()
    active = db.table("crm_deals").select("id", count="exact").eq("deal_status", "ACTIVE").execute()
    inactive = db.table("crm_deals").select("id", count="exact").eq("deal_status", "INACTIVE").execute()
    return {
        "total_customers": customers.count or 0,
        "active_deals": active.count or 0,
        "inactive_deals": inactive.count or 0,
    }

# ── Customers ─────────────────────────────────────────────────────────────────

@router.get("/customers")
def list_customers(
    search: Optional[str] = Query(None),
    provider: Optional[str] = Query(None),
    deal_status: Optional[str] = Query(None),
    meter_type: Optional[str] = Query(None),
    source: Optional[str] = Query(None, description="exact notes value, or __manual__ for none"),
    date_from: Optional[str] = Query(None, description="created on/after YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="created on/before YYYY-MM-DD"),
    limit: int = Query(50),
    offset: int = Query(0),
    user: UserContext = Depends(get_current_user),
):
    db = get_client()
    agent_name = user.sales_agent_name.lower() if user.is_sales_agent and user.sales_agent_name else None

    # Deal-level filters use an inner-join embed so PostgREST filters and
    # paginates server-side (an id pre-fetch breaks past 1000 customers).
    deal_filtered = bool(deal_status or provider or agent_name or meter_type)
    embed = "crm_deals!inner" if deal_filtered else "crm_deals"
    q = db.table("crm_customers").select(
        "id, full_name, first_name, last_name, email, phone, city, state, notes, created_at, "
        f"{embed}(id, deal_status, provider, sales_agent, service_address, business_name)"
    )
    if deal_status:
        q = q.eq("crm_deals.deal_status", deal_status.upper())
    if provider:
        q = q.ilike("crm_deals.provider", provider)
    if agent_name:
        q = q.ilike("crm_deals.sales_agent", agent_name)
    if meter_type:
        q = q.ilike("crm_deals.meter_type", f"%{meter_type}%")
    if search:
        q = q.or_(f"full_name.ilike.%{search}%,email.ilike.%{search}%,phone.ilike.%{search}%")
    if source == "__manual__":
        q = q.is_("notes", "null")
    elif source:
        q = q.eq("notes", source)
    if date_from:
        q = q.gte("created_at", date_from)
    if date_to:
        q = q.lte("created_at", f"{date_to}T23:59:59")
    res = q.order("full_name").range(offset, offset + limit - 1).execute()

    results = []
    for c in res.data:
        deals = c.get("crm_deals", []) or []
        active_count = sum(1 for d in deals if d.get("deal_status") == "ACTIVE")
        active_deals = [d for d in deals if d.get("deal_status") == "ACTIVE"]
        svc_deals = active_deals or deals
        service_address = svc_deals[0].get("service_address") if svc_deals else None
        business_name = next((d.get("business_name") for d in deals if d.get("business_name")), None)
        visible_deals = deals
        if deal_status:
            visible_deals = [d for d in deals if d.get("deal_status", "").upper() == deal_status.upper()]
        prov = next((d.get("provider") for d in (active_deals or deals) if d.get("provider")), None)
        results.append({
            "id": c["id"],
            "full_name": c["full_name"],
            "email": c.get("email"),
            "phone": c.get("phone"),
            "city": c.get("city"),
            "state": c.get("state"),
            "notes": c.get("notes"),
            "source": source_label(c.get("notes")),
            "service_address": service_address,
            "business_name": business_name,
            "provider": prov,
            "deal_count": len(visible_deals),
            "active_deal_count": active_count,
            "created_at": c.get("created_at"),
        })
    return results


def source_label(notes: Optional[str]) -> str:
    """Short human label for where a customer record came from."""
    n = (notes or "").strip()
    if not n:
        return "Manual"
    if n == "HubSpot":
        return "HubSpot"
    if "transferred to Direct Energy" in n:
        return "Direct Energy Transfer"
    if n.endswith("Statement") and len(n) <= 40:
        return n
    return n[:28] + ("…" if len(n) > 28 else "")


@router.get("/customers/sources")
def list_customer_sources(user: UserContext = Depends(get_current_user)):
    """Distinct customer sources with counts, for the list-page filter."""
    db = get_client()
    counts: dict = {}
    off = 0
    while True:
        page = db.table("crm_customers").select("notes").order("id") \
            .range(off, off + 999).execute().data or []
        for c in page:
            n = (c.get("notes") or "").strip()
            counts[n] = counts.get(n, 0) + 1
        if len(page) < 1000:
            break
        off += 1000
    out = []
    for n, cnt in sorted(counts.items(), key=lambda kv: -kv[1]):
        out.append({"value": n if n else "__manual__", "label": source_label(n or None), "count": cnt})
    return out


@router.get("/customers/{id}")
def get_customer(id: str, user: UserContext = Depends(get_current_user)):
    db = get_client()
    c = db.table("crm_customers").select("*").eq("id", id).limit(1).execute()
    if not c.data:
        raise HTTPException(status_code=404, detail="Customer not found")
    deals_res = db.table("crm_deals").select("*").eq("customer_id", id).order("deal_status").execute()
    deals = deals_res.data or []
    _enrich_deals(db, deals)
    # Sales agents only see their own deals for this customer
    if user.is_sales_agent and user.sales_agent_name:
        agent_name = user.sales_agent_name.lower()
        deals = [d for d in deals if (d.get("sales_agent") or "").lower() == agent_name]
        if not deals:
            raise HTTPException(status_code=403, detail="Access denied")
    return {**c.data[0], "deals": deals}

@router.get("/customers/{id}/notes")
def get_customer_notes(id: str, user: UserContext = Depends(get_current_user)):
    db = get_client()
    res = db.table("crm_customer_notes").select("*").eq("crm_customer_id", id).order("created_at", desc=True).execute()
    return res.data or []

@router.post("/customers/{id}/notes")
def create_customer_note(id: str, data: dict = Body(...), user: UserContext = Depends(get_current_user)):
    content = str(data.get("content") or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="Content required")
    author = str(data.get("author_name") or "").strip() or None
    db = get_client()
    res = db.table("crm_customer_notes").insert({"crm_customer_id": id, "content": content, "author_name": author}).execute()
    return res.data[0]

@router.delete("/customers/{id}/notes/{note_id}")
def delete_customer_note(id: str, note_id: str, user: UserContext = Depends(require_admin)):
    db = get_client()
    db.table("crm_customer_notes").delete().eq("id", note_id).eq("crm_customer_id", id).execute()
    return {"ok": True}

# ── Customer attachments ──

ATTACH_BUCKET = "crm-attachments"
MAX_ATTACH_BYTES = 25 * 1024 * 1024  # 25 MB

@router.get("/customers/{id}/attachments")
def list_customer_attachments(id: str, user: UserContext = Depends(get_current_user)):
    db = get_client()
    res = db.table("crm_customer_attachments").select("*").eq("crm_customer_id", id).order("created_at", desc=True).execute()
    return res.data or []

@router.post("/customers/{id}/attachments")
async def upload_customer_attachment(id: str, file: UploadFile = File(...), user: UserContext = Depends(get_current_user)):
    db = get_client()
    customer = db.table("crm_customers").select("id").eq("id", id).limit(1).execute()
    if not customer.data:
        raise HTTPException(status_code=404, detail="Customer not found")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Empty file.")
    if len(file_bytes) > MAX_ATTACH_BYTES:
        raise HTTPException(status_code=400, detail="File too large (max 25 MB).")

    original = file.filename or "upload"
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", original)
    stamp = dt.utcnow().strftime("%Y%m%d%H%M%S%f")
    path = f"{id}/{stamp}_{safe}"
    content_type = file.content_type or "application/octet-stream"

    try:
        db.storage.from_(ATTACH_BUCKET).upload(path, file_bytes, {"content-type": content_type, "upsert": "true"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

    row = {
        "crm_customer_id": id,
        "file_name": original,
        "storage_path": path,
        "file_type": content_type,
        "file_size": len(file_bytes),
        "uploaded_by": user.name or user.email or None,
    }
    res = db.table("crm_customer_attachments").insert(row).execute()
    return res.data[0]

@router.get("/customers/{id}/attachments/{attachment_id}/url")
def get_customer_attachment_url(id: str, attachment_id: str, user: UserContext = Depends(get_current_user)):
    db = get_client()
    res = db.table("crm_customer_attachments").select("storage_path").eq("id", attachment_id).eq("crm_customer_id", id).limit(1).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Attachment not found")
    path = res.data[0]["storage_path"]
    try:
        result = db.storage.from_(ATTACH_BUCKET).create_signed_url(path, 3600)
        if hasattr(result, "data"):
            url = (result.data or {}).get("signedUrl") or (result.data or {}).get("signedURL")
        else:
            url = result.get("signedUrl") or result.get("signedURL")
        if not url:
            raise ValueError("Empty signed URL")
        return {"url": url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate signed URL: {str(e)}")

@router.delete("/customers/{id}/attachments/{attachment_id}")
def delete_customer_attachment(id: str, attachment_id: str, user: UserContext = Depends(get_current_user)):
    db = get_client()
    res = db.table("crm_customer_attachments").select("storage_path").eq("id", attachment_id).eq("crm_customer_id", id).limit(1).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Attachment not found")
    path = res.data[0]["storage_path"]
    try:
        db.storage.from_(ATTACH_BUCKET).remove([path])
    except Exception:
        pass  # row removal still proceeds even if the blob is already gone
    db.table("crm_customer_attachments").delete().eq("id", attachment_id).eq("crm_customer_id", id).execute()
    return {"ok": True}

@router.patch("/customers/{id}")
def update_customer(id: str, data: dict = Body(...), user: UserContext = Depends(get_current_user)):
    db = get_client()
    allowed = {"full_name", "first_name", "last_name", "email", "phone", "dob",
               "mailing_address", "city", "state", "postal_code", "notes"}
    payload = {k: v for k, v in data.items() if k in allowed}
    from datetime import datetime, timezone
    result = {}
    if payload:
        payload["updated_at"] = datetime.now(timezone.utc).isoformat()
        res = db.table("crm_customers").update(payload).eq("id", id).execute()
        result = res.data[0] if res.data else {}
    # ANXH lives on crm_deals — update all deals for this customer
    if "anxh" in data:
        db.table("crm_deals").update({"anxh": data["anxh"] or None}).eq("customer_id", id).execute()
        result["anxh"] = data["anxh"]
    if not payload and "anxh" not in data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    return result

@router.delete("/customers/{id}")
def delete_crm_customer(id: str, user: UserContext = Depends(require_manager)):
    db = get_client()
    # Delete deal notes first (FK: crm_deal_notes → crm_deals)
    deals = db.table("crm_deals").select("id").eq("customer_id", id).execute().data or []
    for deal in deals:
        db.table("crm_deal_notes").delete().eq("crm_deal_id", deal["id"]).execute()
    db.table("crm_deals").delete().eq("customer_id", id).execute()
    db.table("crm_customer_notes").delete().eq("crm_customer_id", id).execute()
    db.table("crm_customers").delete().eq("id", id).execute()
    return {"ok": True}

# ── Deals ─────────────────────────────────────────────────────────────────────

@router.post("/customers/{id}/deals")
def create_customer_deal(id: str, data: dict = Body(...), user: UserContext = Depends(get_current_user)):
    db = get_client()
    c = db.table("crm_customers").select("id").eq("id", id).limit(1).execute()
    if not c.data:
        raise HTTPException(status_code=404, detail="Customer not found")

    def _f(key):
        v = data.get(key)
        try: return float(v) if v not in (None, "", "null") else None
        except (ValueError, TypeError): return None

    payload = {
        "customer_id":           id,
        "deal_name":             str(data.get("deal_name") or "").strip() or None,
        "business_name":         str(data.get("business_name") or "").strip() or None,
        "provider":              str(data.get("provider") or "").strip().upper() or None,
        "esiid":                 str(data.get("esiid") or "").strip() or None,
        "meter_type":            str(data.get("meter_type") or "").strip() or None,
        "deal_type":             str(data.get("deal_type") or "").strip() or None,
        "deal_status":           str(data.get("deal_status") or "ACTIVE").strip().upper(),
        "energy_rate":           _f("energy_rate"),
        "adder":                 _f("adder"),
        "contract_term":         str(data.get("contract_term") or "").strip() or None,
        "contract_start_date":   data.get("contract_start_date") or None,
        "contract_end_date":     data.get("contract_end_date") or None,
        "contract_signed_date":  data.get("contract_signed_date") or None,
        "service_address":       str(data.get("service_address") or "").strip() or None,
        "sales_agent":           str(data.get("sales_agent") or "").strip() or None,
        "deal_owner":            str(data.get("deal_owner") or "").strip() or None,
        "product_type":          str(data.get("product_type") or "").strip() or None,
        "flag_tos":              bool(data.get("flag_tos", False)),
        "flag_toao":             bool(data.get("flag_toao", False)),
        "flag_deposit":          bool(data.get("flag_deposit", False)),
        "flag_special_deal":     bool(data.get("flag_special_deal", False)),
        "flag_promo_10":         bool(data.get("flag_promo_10", False)),
        "flag_delinked":         bool(data.get("flag_delinked", False)),
    }
    res = db.table("crm_deals").insert(payload).execute()
    if not res.data:
        raise HTTPException(status_code=500, detail="Failed to create deal")
    return res.data[0]

@router.get("/deals")
def list_deals(
    search: Optional[str] = Query(None),
    provider: Optional[str] = Query(None),
    deal_status: Optional[str] = Query(None),
    meter_type: Optional[str] = Query(None),
    deal_type: Optional[str] = Query(None),
    sales_agent: Optional[str] = Query(None),
    limit: int = Query(50),
    offset: int = Query(0),
    user: UserContext = Depends(get_current_user),
):
    db = get_client()
    q = db.table("crm_deals").select(
        "*, crm_customers(id, full_name, email)"
    )
    if provider:
        q = q.ilike("provider", f"%{provider}%")
    if deal_status:
        q = q.eq("deal_status", deal_status.upper())
    if meter_type:
        q = q.ilike("meter_type", f"%{meter_type}%")
    if deal_type:
        q = q.ilike("deal_type", f"%{deal_type}%")
    if sales_agent:
        q = q.ilike("sales_agent", f"%{sales_agent}%")
    if search:
        q = q.or_(f"deal_name.ilike.%{search}%,esiid.ilike.%{search}%,service_address.ilike.%{search}%")
    res = q.order("created_at", desc=True).range(offset, offset + limit - 1).execute()
    deals = res.data or []
    _enrich_deals(db, deals)
    # Sales agents only see their own deals
    if user.is_sales_agent and user.sales_agent_name:
        deals = [d for d in deals if (d.get("sales_agent") or "").lower() == user.sales_agent_name.lower()]
    return deals

@router.get("/deals/check-duplicate")
def check_duplicate_deal(
    esiid: Optional[str] = Query(None),
    service_address: Optional[str] = Query(None),
    active_only: bool = Query(False),
    user: UserContext = Depends(get_current_user),
):
    db = get_client()
    matches = []
    seen_ids: set = set()

    def _add_crm(rows):
        for d in (rows or []):
            if d["id"] in seen_ids:
                continue
            seen_ids.add(d["id"])
            matches.append({
                "source": "crm",
                "deal_id": d["id"],
                "deal_name": d.get("deal_name") or "Unnamed Deal",
                "provider": d.get("provider") or "",
                "deal_status": d.get("deal_status") or "",
                "esiid": d.get("esiid") or "",
                "service_address": d.get("service_address") or "",
                "customer_name": (d.get("crm_customers") or {}).get("full_name") or "",
                "customer_id": (d.get("crm_customers") or {}).get("id") or "",
            })

    def _add_lead(rows):
        for d in (rows or []):
            if d["id"] in seen_ids:
                continue
            seen_ids.add(d["id"])
            lead = d.get("crm_leads") or {}
            matches.append({
                "source": "lead",
                "deal_id": d["id"],
                "deal_name": d.get("plan_name") or "Unnamed Deal",
                "provider": d.get("supplier") or "",
                "deal_status": d.get("status") or "",
                "esiid": d.get("esiid") or "",
                "service_address": d.get("service_address") or "",
                "customer_name": f"{lead.get('first_name','')} {lead.get('last_name','')}".strip(),
                "customer_id": lead.get("id") or "",
            })

    if esiid and esiid.strip():
        esiid_clean = esiid.strip()
        crm_q = db.table("crm_deals").select(
            "id, deal_name, provider, deal_status, esiid, service_address, crm_customers(id, full_name)"
        ).eq("esiid", esiid_clean)
        if active_only:
            crm_q = crm_q.eq("deal_status", "ACTIVE")
        _add_crm(crm_q.execute().data)

        lead_q = db.table("lead_deals").select(
            "id, plan_name, supplier, status, esiid, service_address, crm_leads(id, first_name, last_name)"
        ).eq("esiid", esiid_clean)
        if active_only:
            lead_q = lead_q.eq("status", "Active")
        _add_lead(lead_q.execute().data)

    if service_address and service_address.strip():
        addr_clean = service_address.strip()
        crm_q = db.table("crm_deals").select(
            "id, deal_name, provider, deal_status, esiid, service_address, crm_customers(id, full_name)"
        ).ilike("service_address", f"%{addr_clean}%")
        if active_only:
            crm_q = crm_q.eq("deal_status", "ACTIVE")
        _add_crm(crm_q.execute().data)

        lead_q = db.table("lead_deals").select(
            "id, plan_name, supplier, status, esiid, service_address, crm_leads(id, first_name, last_name)"
        ).ilike("service_address", f"%{addr_clean}%")
        if active_only:
            lead_q = lead_q.eq("status", "Active")
        _add_lead(lead_q.execute().data)

    return {"matches": matches}

@router.get("/deals/{id}")
def get_deal(id: str, user: UserContext = Depends(get_current_user)):
    db = get_client()
    res = db.table("crm_deals").select("*, crm_customers(id, full_name, email, phone)").eq("id", id).limit(1).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Deal not found")
    deal = res.data[0]
    _enrich_deals(db, [deal])
    if user.is_sales_agent and user.sales_agent_name:
        if (deal.get("sales_agent") or "").lower() != user.sales_agent_name.lower():
            raise HTTPException(status_code=403, detail="Access denied")
    return deal

@router.get("/deals/{id}/notes")
def get_deal_notes(id: str, user: UserContext = Depends(get_current_user)):
    db = get_client()
    res = db.table("crm_deal_notes").select("*").eq("crm_deal_id", id).order("created_at", desc=True).execute()
    return res.data or []

@router.post("/deals/{id}/notes")
def create_deal_note(id: str, data: dict = Body(...), user: UserContext = Depends(get_current_user)):
    content = str(data.get("content") or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="Content required")
    author = str(data.get("author_name") or "").strip() or None
    db = get_client()
    res = db.table("crm_deal_notes").insert({"crm_deal_id": id, "content": content, "author_name": author}).execute()
    return res.data[0]

@router.post("/deals/{id}/renew")
def renew_deal(id: str, data: dict = Body(...), user: UserContext = Depends(get_current_user)):
    db = get_client()
    orig_res = db.table("crm_deals").select("*").eq("id", id).limit(1).execute()
    if not orig_res.data:
        raise HTTPException(status_code=404, detail="Deal not found")
    orig = orig_res.data[0]

    # Mark original deal as RENEWED
    db.table("crm_deals").update({"deal_status": "RENEWED"}).eq("id", id).execute()

    # Build new deal — inherit ESIID/address/meter from original, override with form data
    new_deal = {
        "customer_id":          orig["customer_id"],
        "esiid":                orig.get("esiid"),
        "provider":             data.get("provider") or orig.get("provider"),
        "deal_status":          "ACTIVE",
        "sales_agent":          data.get("sales_agent") or orig.get("sales_agent"),
        "deal_owner":           data.get("deal_owner") or orig.get("deal_owner"),
        "service_address":      orig.get("service_address"),
        "contract_start_date":  data.get("contract_start_date") or None,
        "contract_end_date":    data.get("contract_end_date") or None,
        "contract_signed_date": data.get("contract_signed_date") or None,
        "contract_term":        data.get("contract_term") or orig.get("contract_term"),
        "energy_rate":          data.get("energy_rate") or None,
        "adder":                data.get("adder") or None,
        "product_type":         data.get("product_type") or orig.get("product_type"),
        "meter_type":           data.get("meter_type") or orig.get("meter_type"),
        "anxh":                 orig.get("anxh"),
        "business_name":        orig.get("business_name"),
        "deal_name":            data.get("deal_name") or orig.get("deal_name"),
    }
    res = db.table("crm_deals").insert(new_deal).execute()
    return res.data[0]

@router.delete("/deals/{id}")
def delete_deal(id: str, user: UserContext = Depends(require_manager)):
    db = get_client()
    db.table("crm_deal_notes").delete().eq("crm_deal_id", id).execute()
    db.table("crm_deals").delete().eq("id", id).execute()
    return {"ok": True}

@router.delete("/deals/{id}/notes/{note_id}")
def delete_deal_note(id: str, note_id: str, user: UserContext = Depends(require_admin)):
    db = get_client()
    db.table("crm_deal_notes").delete().eq("id", note_id).eq("crm_deal_id", id).execute()
    return {"ok": True}

@router.patch("/deals/{id}")
def update_deal(id: str, data: dict = Body(...), user: UserContext = Depends(get_current_user)):
    db = get_client()
    allowed = {"deal_status", "deal_name", "provider", "adder", "energy_rate", "deal_owner",
               "sales_agent", "contract_start_date", "contract_end_date", "contract_signed_date",
               "contract_term", "notes", "service_address", "meter_type", "deal_type",
               "business_name", "anxh", "esiid", "product_type",
               "flag_tos", "flag_toao", "flag_deposit", "flag_special_deal", "flag_promo_10", "flag_delinked"}
    payload = {k: v for k, v in data.items() if k in allowed}
    if not payload:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    from datetime import datetime, timezone
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = db.table("crm_deals").update(payload).eq("id", id).execute()
    return res.data[0] if res.data else {}

# ── Import ─────────────────────────────────────────────────────────────────────

@router.post("/import")
def import_deals(file_path: str = Body(..., embed=True), user: UserContext = Depends(require_admin)):
    """
    One-time idempotent import from MERGED_DEALS FINAL.xlsx.
    Deduplicates customers by email (from Associated Contact column).
    """
    import openpyxl

    db = get_client()

    # Load supplier code → id map
    sup_res = db.table("suppliers").select("id, code").execute()
    supplier_map = {s["code"]: s["id"] for s in sup_res.data}

    # Load existing customers by email (idempotency)
    existing_res = db.table("crm_customers").select("id, email").execute()
    customer_by_email: dict = {c["email"]: c["id"] for c in existing_res.data if c.get("email")}

    # Load existing ESIIDs to prevent duplicate deals
    esiid_res = db.table("crm_deals").select("esiid").execute()
    existing_esiids = {r["esiid"] for r in (esiid_res.data or []) if r.get("esiid")}

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)  # skip header

    customers_created = 0
    deals_created = 0
    deals_skipped = 0

    for row in rows_iter:
        if len(row) < 20:
            continue

        adder        = _to_float(row[0])
        signed_date  = _to_date_str(row[1])
        end_date     = _to_date_str(row[2])
        rate         = _to_float(row[3])
        start_date   = _to_date_str(row[4])
        term         = str(row[5] or "").strip() or None
        deal_name    = str(row[6] or "").strip() or None
        biz_name     = str(row[7] or "").strip() or None
        deal_owner   = str(row[8] or "").strip() or None
        deal_status  = str(row[9] or "ACTIVE").strip().upper() or "ACTIVE"
        meter_type   = str(row[10] or "").strip() or None
        deal_type    = str(row[11] or "").strip() or None
        esiid        = str(row[12] or "").strip() or None
        product_type = str(row[13] or "").strip() or None
        sales_agent  = str(row[14] or "").strip() or None
        svc_address  = str(row[15] or "").strip() or None
        provider     = str(row[16] or "").strip().upper() or None
        contact      = str(row[17] or "").strip()
        first_name   = str(row[18] or "").strip() or None
        last_name    = str(row[19] or "").strip() or None
        anxh         = str(row[20] or "").strip() or None
        dob          = str(row[21] or "").strip() or None
        email        = str(row[22] or "").strip().lower() or None
        mail_addr    = str(row[23] or "").strip() or None
        city         = str(row[24] or "").strip() or None
        postal       = str(row[25] or "").strip() or None
        state        = str(row[26] or "TX").strip() or "TX"
        phone        = str(row[28] or "").strip() if len(row) > 28 else None

        # Try email from col 22 first, fallback to contact column
        if not email:
            email = _extract_email(contact)

        # Build customer key
        if email:
            cust_key = email
        elif first_name or last_name:
            cust_key = f"{first_name or ''} {last_name or ''}".strip().lower()
        else:
            deals_skipped += 1
            continue

        # Get or create customer
        if cust_key not in customer_by_email:
            full_name = f"{first_name or ''} {last_name or ''}".strip()
            if not full_name:
                # Try extracting name from contact string
                m = re.match(r"^([^(]+)\s*\(", contact)
                full_name = m.group(1).strip() if m else cust_key

            new_cust = {
                "full_name": full_name,
                "first_name": first_name,
                "last_name": last_name,
                "email": email,
                "phone": phone,
                "dob": dob,
                "mailing_address": mail_addr,
                "city": city,
                "state": state[:2] if state else "TX",
                "postal_code": postal,
            }
            cres = db.table("crm_customers").insert(new_cust).execute()
            customer_id = cres.data[0]["id"]
            customer_by_email[cust_key] = customer_id
            customers_created += 1
        else:
            customer_id = customer_by_email[cust_key]

        # Map provider to supplier_id
        supplier_code = PROVIDER_TO_CODE.get(provider or "")
        supplier_id = supplier_map.get(supplier_code) if supplier_code else None

        # Flag truncated ESIIDs (scientific notation)
        esiid_notes = "ESIID truncated (scientific notation in source file)" if esiid and ("E+" in esiid or "e+" in esiid) else None

        deal = {
            "customer_id": customer_id,
            "deal_name": deal_name,
            "business_name": biz_name,
            "esiid": esiid,
            "provider": provider,
            "supplier_id": supplier_id,
            "meter_type": meter_type,
            "deal_type": deal_type,
            "deal_status": deal_status if deal_status in ("ACTIVE", "INACTIVE") else "ACTIVE",
            "adder": adder,
            "energy_rate": rate,
            "product_type": product_type,
            "contract_term": term,
            "contract_signed_date": signed_date,
            "contract_start_date": start_date,
            "contract_end_date": end_date,
            "service_address": svc_address,
            "deal_owner": deal_owner,
            "sales_agent": sales_agent,
            "anxh": anxh,
        }
        if esiid and esiid in existing_esiids:
            deals_skipped += 1
            continue
        db.table("crm_deals").insert(deal).execute()
        if esiid:
            existing_esiids.add(esiid)
        deals_created += 1

    wb.close()
    return {
        "customers_created": customers_created,
        "deals_created": deals_created,
        "deals_skipped": deals_skipped,
    }

# ── Providers list (for filter dropdowns) ─────────────────────────────────────

def _paginate_distinct(db, table: str, column: str) -> list:
    """Fetch all distinct non-null values for a column, paginating past the 1000-row limit."""
    values: set = set()
    offset = 0
    while True:
        batch = db.table(table).select(column).range(offset, offset + 999).execute().data or []
        for r in batch:
            v = r.get(column)
            if v:
                values.add(v)
        if len(batch) < 1000:
            break
        offset += 1000
    return sorted(values)

@router.get("/providers")
def list_providers(user: UserContext = Depends(get_current_user)):
    db = get_client()
    return _paginate_distinct(db, "crm_deals", "provider")

@router.get("/agents")
def list_agents(user: UserContext = Depends(get_current_user)):
    db = get_client()
    return _paginate_distinct(db, "crm_deals", "sales_agent")


# ── Import Template Download ───────────────────────────────────────────────────

@router.get("/import-template")
def download_import_template(user: UserContext = Depends(require_admin)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"

    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor="0F1D5E")
    sample_font   = Font(italic=True, color="888888")
    note_font     = Font(bold=True, color="D97706")

    # Row 1: headers
    for col, h in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Row 2: sample data
    for col, val in enumerate(SAMPLE_ROW, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = sample_font

    # Row 4: valid suppliers note
    note_cell = ws.cell(row=4, column=1, value="✦ Valid Supplier Names:")
    note_cell.font = note_font
    for col, sup in enumerate(VALID_SUPPLIERS, 2):
        ws.cell(row=4, column=col, value=sup).font = Font(color="666666", italic=True)

    # Row 5: required fields note
    req_cell = ws.cell(row=5, column=1, value="✦ Required fields: First Name, Last Name, ESIID, Supplier")
    req_cell.font = Font(color="DC2626", bold=True)

    # Column widths
    col_widths = [12, 12, 28, 14, 14, 28, 14, 8, 10,
                  22, 18, 12, 18, 28, 18, 18, 20,
                  12, 12, 10, 16, 14, 20, 20, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="crm_import_template.xlsx"'},
    )


# ── Import Upload ──────────────────────────────────────────────────────────────

def _norm_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ("nan", "none", "nat") else None

def _norm_date(val) -> Optional[str]:
    if not val:
        return None
    s = _norm_str(val)
    if not s:
        return None
    if hasattr(val, "date"):
        return val.date().isoformat()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%-m/%-d/%Y", "%m-%d-%Y"):
        try:
            return dt.strptime(s[:10], fmt).date().isoformat()
        except Exception:
            continue
    return None

def _norm_float(val) -> Optional[float]:
    try:
        s = str(val).replace(",", "").strip()
        return float(s) if s and s.lower() not in ("nan", "none", "") else None
    except Exception:
        return None

def _norm_supplier(val) -> Optional[str]:
    if not val:
        return None
    return SUPPLIER_ALIASES.get(str(val).strip().lower())

def _norm_status(val) -> str:
    s = str(val or "").strip().lower()
    if s in INACTIVE_STATUSES:
        return "INACTIVE"
    return "ACTIVE"

@router.post("/import-upload")
async def import_upload(
    file: UploadFile = File(...),
    user: UserContext = Depends(require_admin),
):
    import openpyxl
    import pandas as pd

    contents = await file.read()
    buf = io.BytesIO(contents)

    # Parse file
    fname = (file.filename or "").lower()
    try:
        if fname.endswith(".csv"):
            df = pd.read_csv(buf)
        else:
            df = pd.read_excel(buf, engine="openpyxl")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {e}")

    # Normalize column names: strip, lowercase for lookup
    col_map = {c.strip().lower(): c for c in df.columns}

    def get_col(row, *names):
        for n in names:
            c = col_map.get(n.lower())
            if c and c in row.index:
                return row[c]
        return None

    db = get_client()

    # Load supplier code → id
    sup_res = db.table("suppliers").select("id, code").execute()
    supplier_id_map = {s["code"]: s["id"] for s in (sup_res.data or [])}

    # Load existing customer emails for dedup
    existing_res = db.table("crm_customers").select("id, email, first_name, last_name").execute()
    customer_by_key: dict = {}
    for c in (existing_res.data or []):
        if c.get("email"):
            customer_by_key[c["email"].lower()] = c["id"]
        else:
            name_key = f"{c.get('first_name','')} {c.get('last_name','')}".strip().lower()
            if name_key:
                customer_by_key[name_key] = c["id"]

    # Load existing ESIIDs for dedup
    esiid_res = db.table("crm_deals").select("esiid").execute()
    existing_esiids = {r["esiid"] for r in (esiid_res.data or []) if r.get("esiid")}

    customers_created = 0
    customers_reused  = 0
    deals_created     = 0
    deals_skipped     = 0
    errors: list      = []

    for idx, row in df.iterrows():
        row_num = idx + 2  # 1-indexed + header row

        first_name = _norm_str(get_col(row, "First Name", "firstname"))
        last_name  = _norm_str(get_col(row, "Last Name", "lastname"))
        if not first_name and not last_name:
            errors.append(f"Row {row_num}: skipped — missing First Name and Last Name")
            deals_skipped += 1
            continue

        esiid = _norm_str(get_col(row, "ESIID", "ESI ID", "esi_id"))
        if not esiid:
            errors.append(f"Row {row_num}: skipped — missing ESIID")
            deals_skipped += 1
            continue

        email = _norm_str(get_col(row, "Email", "email address"))
        email = email.lower() if email else None

        # Customer dedup key
        if email:
            cust_key = email
        else:
            cust_key = f"{first_name or ''} {last_name or ''}".strip().lower()

        # Get or create customer
        if cust_key in customer_by_key:
            customer_id = customer_by_key[cust_key]
            customers_reused += 1
        else:
            full_name = f"{first_name or ''} {last_name or ''}".strip()
            state_val = _norm_str(get_col(row, "State")) or "TX"
            new_cust = {
                "full_name":       full_name,
                "first_name":      first_name,
                "last_name":       last_name,
                "email":           email,
                "phone":           _norm_str(get_col(row, "Phone")),
                "dob":             _norm_str(get_col(row, "Date of Birth", "DOB")),
                "mailing_address": _norm_str(get_col(row, "Mailing Address", "Address")),
                "city":            _norm_str(get_col(row, "City")),
                "state":           state_val[:2].upper() if state_val else "TX",
                "postal_code":     _norm_str(get_col(row, "Zip Code", "Postal Code", "ZIP")),
            }
            cres = db.table("crm_customers").insert(new_cust).execute()
            customer_id = cres.data[0]["id"]
            customer_by_key[cust_key] = customer_id
            customers_created += 1

        # ESIID dedup
        if esiid in existing_esiids:
            deals_skipped += 1
            continue

        # Supplier
        supplier_raw  = _norm_str(get_col(row, "Supplier", "Provider"))
        supplier_code = _norm_supplier(supplier_raw)
        supplier_id   = supplier_id_map.get(supplier_code) if supplier_code else None
        provider_name = supplier_raw.upper() if supplier_raw else None

        deal = {
            "customer_id":          customer_id,
            "esiid":                esiid,
            "provider":             provider_name,
            "supplier_id":          supplier_id,
            "deal_status":          _norm_status(get_col(row, "Deal Status", "Status")),
            "sales_agent":          _norm_str(get_col(row, "Sales Agent")),
            "deal_owner":           _norm_str(get_col(row, "Deal Owner")),
            "service_address":      _norm_str(get_col(row, "Service Address")),
            "contract_start_date":  _norm_date(get_col(row, "Contract Start Date", "Start Date")),
            "contract_end_date":    _norm_date(get_col(row, "Contract End Date", "End Date")),
            "contract_signed_date": _norm_date(get_col(row, "Contract Signed Date", "Signed Date")),
            "contract_term":        _norm_str(get_col(row, "Term (Months)", "Term")),
            "energy_rate":          _norm_float(get_col(row, "Energy Rate", "Rate")),
            "adder":                _norm_float(get_col(row, "Adder")),
            "product_type":         _norm_str(get_col(row, "Product Type")),
            "meter_type":           _norm_str(get_col(row, "Meter Type")),
            "anxh":                 _norm_str(get_col(row, "ANXH", "anxh")),
            "business_name":        _norm_str(get_col(row, "Business Name")),
            "deal_name":            _norm_str(get_col(row, "Deal Name")),
        }
        db.table("crm_deals").insert(deal).execute()
        existing_esiids.add(esiid)
        deals_created += 1

    return {
        "customers_created": customers_created,
        "customers_reused":  customers_reused,
        "deals_created":     deals_created,
        "deals_skipped":     deals_skipped,
        "errors":            errors,
    }


# ── Clear All Imported Data ────────────────────────────────────────────────────

@router.delete("/clear")
def clear_crm_data(user: UserContext = Depends(require_admin)):
    db = get_client()

    # Count before delete for reporting
    cust_count = db.table("crm_customers").select("id", count="exact").execute().count or 0
    deal_count = db.table("crm_deals").select("id", count="exact").execute().count or 0

    # Wipe in FK order
    deal_ids = [r["id"] for r in (db.table("crm_deals").select("id").execute().data or [])]
    if deal_ids:
        db.table("crm_deal_notes").delete().in_("crm_deal_id", deal_ids).execute()
    db.table("crm_deals").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()

    cust_ids = [r["id"] for r in (db.table("crm_customers").select("id").execute().data or [])]
    if cust_ids:
        db.table("crm_customer_notes").delete().in_("crm_customer_id", cust_ids).execute()
    db.table("crm_customers").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()

    return {"deleted_customers": cust_count, "deleted_deals": deal_count}


# ── Deduplicate deals by ESIID ─────────────────────────────────────────────────

@router.post("/deduplicate-deals")
def deduplicate_deals(user: UserContext = Depends(require_admin)):
    """Remove duplicate crm_deals rows, keeping the earliest (lowest created_at) per ESIID."""
    db = get_client()

    all_deals = db.table("crm_deals").select("id, esiid, created_at").order("created_at").execute().data or []

    seen_esiids: set = set()
    to_delete: list = []
    for d in all_deals:
        esiid = d.get("esiid")
        if not esiid:
            continue
        if esiid in seen_esiids:
            to_delete.append(d["id"])
        else:
            seen_esiids.add(esiid)

    if to_delete:
        db.table("crm_deal_notes").delete().in_("crm_deal_id", to_delete).execute()
        db.table("crm_deals").delete().in_("id", to_delete).execute()

    return {"duplicates_removed": len(to_delete)}
